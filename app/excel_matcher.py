from __future__ import annotations

from pathlib import Path
from rapidfuzz import fuzz, process
import openpyxl

from .models import MasterRow
from .utils import normalize_name, normalize_spec


REGION_COLUMNS = {
    '수도권': {'regular_price': 17, 'kg_price': 18, 'unit_price': 19},
    '경상권': {'regular_price': 20, 'kg_price': None, 'unit_price': 21},
    '경남부산': {'regular_price': 22, 'kg_price': None, 'unit_price': 23},
    '호남권': {'regular_price': 24, 'kg_price': None, 'unit_price': 25},
}

REGION_ALIASES = {
    '수도권': '수도권',
    '경상권': '경상권',
    '경남 부산': '경남부산',
    '경남부산': '경남부산',
    '호남권': '호남권',
}


class MasterCatalog:
    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self.rows: list[MasterRow] = []
        self._choices: dict[str, MasterRow] = {}
        self._load()

    def _load(self) -> None:
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row_idx in range(4, ws.max_row + 1):
            code = ws.cell(row_idx, 1).value
            name = ws.cell(row_idx, 3).value
            spec = ws.cell(row_idx, 4).value
            if not name or not spec:
                continue
            prices: dict[str, dict[str, float | None]] = {}
            for region, cols in REGION_COLUMNS.items():
                region_prices = {}
                for key, col in cols.items():
                    val = ws.cell(row_idx, col).value if col else None
                    if val in (None, '', '-'):
                        region_prices[key] = None
                    else:
                        try:
                            region_prices[key] = float(val)
                        except Exception:
                            region_prices[key] = None
                prices[region] = region_prices
            master = MasterRow(
                code=str(code or ''),
                name=str(name).strip(),
                spec=str(spec).strip(),
                normalized_name=normalize_name(str(name)),
                normalized_spec=normalize_spec(str(spec)),
                prices=prices,
            )
            self.rows.append(master)
            self._choices[master.normalized_name] = master

    def find_best_match(self, product_name: str, spec: str = '') -> tuple[MasterRow | None, float]:
        q_name = normalize_name(product_name)
        q_spec = normalize_spec(spec)
        if not q_name:
            return None, 0.0
        best_row = None
        best_score = 0.0
        for row in self.rows:
            name_score = fuzz.WRatio(q_name, row.normalized_name)
            spec_score = fuzz.WRatio(q_spec, row.normalized_spec) if q_spec else 0
            combined = name_score * 0.82 + spec_score * 0.18
            if combined > best_score:
                best_score = combined
                best_row = row
        return best_row, round(best_score, 1)
