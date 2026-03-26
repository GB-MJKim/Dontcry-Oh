import json
import os
import re
import shutil
from typing import Dict, List, Optional, Tuple

import openpyxl
from rapidfuzz import fuzz

from .settings import MASTER_META_PATH, MASTER_XLSX_PATH

REGION_SU = "\uC218\uB3C4\uAD8C"
REGION_GS = "\uACBD\uC0C1\uAD8C"
REGION_GN = "\uACBD\uB0A8\uBD80\uC0B0"
REGION_HN = "\uD638\uB0A8\uAD8C"

EXCEL_LAYOUT = {
    "start_row": 4,
    "code": 1,
    "name": 3,
    "spec": 4,
    "regions": {
        REGION_SU: {"spec": 17, "kg": 18, "unit": 19},
        REGION_GS: {"spec": 20, "kg": None, "unit": 21},
        REGION_GN: {"spec": 22, "kg": None, "unit": 23},
        REGION_HN: {"spec": 24, "kg": None, "unit": 25},
    },
}

NAME_MATCH_THRESHOLD = 58.0
REMOVE_BRACKET_WORDS = [
    "\uB0C9\uB3D9",
    "\uB0C9\uC7A5",
    "\uC0C1\uC628",
    "\uC0DD\uC0B0",
    "\uBC18\uC81C",
    "\uC644\uC81C",
    "\uACC4\uC808\uC0C1\uD488",
    "\uC2E0\uC0C1\uD488",
]
EXCLUDE_KEYWORDS = [
    "\uC99D\uC815",
    "\uC0AC\uC740\uD488",
    "\uC138\uD2B8\uC0C1\uD488",
    "\uC138\uD2B8",
    "\uBB34\uB8CC",
]

CATALOG_CACHE: Optional[List[dict]] = None


def _safe_str(value) -> str:
    return "" if value is None else str(value).strip()


def _extract_date_from_filename(filename: str) -> Optional[dict]:
    raw = _safe_str(filename)
    if not raw:
        return None

    match = re.search(
        "(?P<year>\\d{2,4})\uB144\\s*(?P<month>\\d{1,2})\uC6D4\\s*(?P<day>\\d{1,2})\uC77C",
        raw,
    )
    if not match:
        return None

    year = int(match.group("year"))
    month = int(match.group("month"))
    day = int(match.group("day"))
    if year < 100:
        year += 2000

    return {
        "iso_date": f"{year:04d}-{month:02d}-{day:02d}",
        "display_date": f"{year:04d}\uB144 {month:02d}\uC6D4 {day:02d}\uC77C",
        "raw_date": match.group(0),
    }


def load_master_metadata() -> Dict[str, Optional[str]]:
    default = {
        "source_filename": None,
        "source_date": None,
        "source_date_iso": None,
        "source_date_raw": None,
    }
    if not os.path.exists(MASTER_META_PATH):
        return default
    try:
        with open(MASTER_META_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return {
            "source_filename": data.get("source_filename"),
            "source_date": data.get("source_date"),
            "source_date_iso": data.get("source_date_iso"),
            "source_date_raw": data.get("source_date_raw"),
        }
    except Exception:
        return default


def save_master_excel(src_path: str, original_filename: Optional[str] = None) -> None:
    os.makedirs(os.path.dirname(str(MASTER_XLSX_PATH)), exist_ok=True)
    shutil.copyfile(src_path, MASTER_XLSX_PATH)

    metadata = {
        "source_filename": _safe_str(original_filename) or os.path.basename(src_path),
        "source_date": None,
        "source_date_iso": None,
        "source_date_raw": None,
    }
    date_info = _extract_date_from_filename(original_filename or "")
    if date_info:
        metadata["source_date"] = date_info["display_date"]
        metadata["source_date_iso"] = date_info["iso_date"]
        metadata["source_date_raw"] = date_info["raw_date"]

    with open(MASTER_META_PATH, "w", encoding="utf-8") as handle:
        json.dump(metadata, handle, ensure_ascii=False, indent=2)

    clear_cache()


def clear_cache() -> None:
    global CATALOG_CACHE
    CATALOG_CACHE = None


def _to_number(value) -> Optional[int]:
    if value in (None, "", "-", "None"):
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = str(value).strip().replace(",", "").replace("\uC6D0", "").replace(" ", "")
    if not text or text.lower() == "nan" or text == "-":
        return None
    try:
        return int(round(float(text)))
    except Exception:
        digits = "".join(ch for ch in text if ch.isdigit())
        return int(digits) if digits else None


def _normalize_text(text: str) -> str:
    value = _safe_str(text).replace("\n", " ").replace("\r", " ")
    value = value.replace("\uD6DE", "x").replace("\u00D7", "x").replace("\uFF0A", "*")
    value = value.replace("\u2013", "-").replace("\u2014", "-").replace("~", "-")
    value = re.sub(r"\s+", " ", value).strip().lower()
    return value


def _clean_brackets_for_name(text: str) -> str:
    value = _safe_str(text)

    def repl(match: re.Match) -> str:
        inner = _normalize_text(match.group(1)).replace(" ", "")
        for word in REMOVE_BRACKET_WORDS:
            if word.replace(" ", "") in inner:
                return " "
        return f" {match.group(1)} "

    return re.sub(r"\((.*?)\)", repl, value)


def _normalize_name(text: str) -> str:
    value = _clean_brackets_for_name(text)
    value = _normalize_text(value)
    value = value.replace("/", " ").replace("_", " ").replace("-", " ")
    value = value.replace(":", " ").replace(",", " ").replace("*", " ")
    value = re.sub("\uBD80\uC7AC\uB8CC.*", " ", value)
    value = re.sub("\uC911\uB7C9.*", " ", value)
    value = re.sub(r"\d+\s*%", " ", value)
    value = re.sub(r"[^0-9a-z\uAC00-\uD7A3]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value.replace(" ", "")


def _normalize_spec(text: str) -> str:
    value = _normalize_text(text)
    value = re.sub(r"\s+", "", value)
    value = value.replace("g\u00D7", "gx").replace("kg\u00D7", "kgx")
    value = value.replace("\uAC1C\uC785", "\uAC1C").replace("\uC774\uC678", "")
    return value


def _extract_numeric_tokens(text: str) -> List[str]:
    return re.findall(r"\d+(?:\.\d+)?(?:kg|g|ml|l|\uAC1C|x)?", _normalize_spec(text))


def _spec_similarity(a: str, b: str) -> float:
    a_norm = _normalize_spec(a)
    b_norm = _normalize_spec(b)
    if not a_norm or not b_norm:
        return 0.0

    score = max(
        fuzz.WRatio(a_norm, b_norm),
        fuzz.partial_ratio(a_norm, b_norm),
        fuzz.token_sort_ratio(a_norm, b_norm),
    )
    a_tokens = set(_extract_numeric_tokens(a_norm))
    b_tokens = set(_extract_numeric_tokens(b_norm))
    if a_tokens and b_tokens:
        inter = len(a_tokens & b_tokens)
        union = len(a_tokens | b_tokens)
        if union:
            score = max(score, 100.0 * inter / union)
    return float(score)


def _expand_variants_from_bracket(base: str, inner: str) -> List[str]:
    out: List[str] = [f"{base} {inner}"]
    if ":" in inner:
        left, right = inner.split(":", 1)
        left = left.strip()
        right_parts = [item.strip() for item in right.split("/") if item.strip()]
        if left:
            out.append(f"{base} {left}")
        for part in right_parts:
            out.append(f"{base} {part}")
            if left:
                out.append(f"{base} {left} {part}")
    else:
        for part in [item.strip() for item in inner.split("/") if item.strip()]:
            out.append(f"{base} {part}")
    return out


def name_candidates(text: str) -> List[str]:
    raw = _safe_str(text).replace("\n", " ").strip()
    if not raw:
        return []

    candidates: List[str] = [raw]
    match = re.search(r"^(.*?)\((.*?)\)\s*$", raw)
    if match:
        base = match.group(1).strip()
        inner = match.group(2).strip()
        candidates.append(base)
        candidates.extend(_expand_variants_from_bracket(base, inner))
    candidates.append(re.sub(r"\(.*?\)", " ", raw))
    candidates.extend([item.strip() for item in re.split(r"[\/]", raw) if item.strip()])

    normalized: List[str] = []
    seen = set()
    for candidate in candidates:
        norm = _normalize_name(candidate)
        if norm and norm not in seen:
            seen.add(norm)
            normalized.append(norm)
    return normalized


def normalize_name(text: str) -> str:
    return _normalize_name(text)


def normalize_spec(text: str) -> str:
    return _normalize_spec(text)


def _containment_bonus(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 100.0
    shorter = min(len(a), len(b))
    if shorter >= 6 and (a in b or b in a):
        return 95.0
    return 0.0


def load_master_df() -> List[dict]:
    global CATALOG_CACHE
    if CATALOG_CACHE is not None:
        return CATALOG_CACHE
    if not os.path.exists(MASTER_XLSX_PATH):
        CATALOG_CACHE = []
        return CATALOG_CACHE

    workbook = openpyxl.load_workbook(MASTER_XLSX_PATH, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: List[dict] = []
    for row_idx in range(EXCEL_LAYOUT["start_row"], sheet.max_row + 1):
        name = _safe_str(sheet.cell(row_idx, EXCEL_LAYOUT["name"]).value)
        spec = _safe_str(sheet.cell(row_idx, EXCEL_LAYOUT["spec"]).value)
        if not name or not spec:
            continue

        row = {
            "row_index": row_idx,
            "code": _safe_str(sheet.cell(row_idx, EXCEL_LAYOUT["code"]).value),
            "name": name,
            "spec": spec,
            "normalized_name": _normalize_name(name),
            "normalized_spec": _normalize_spec(spec),
            "name_candidates": name_candidates(name),
        }
        for region, cols in EXCEL_LAYOUT["regions"].items():
            row[f"{region}_spec_price"] = _to_number(sheet.cell(row_idx, cols["spec"]).value) if cols["spec"] else None
            row[f"{region}_kg_price"] = _to_number(sheet.cell(row_idx, cols["kg"]).value) if cols["kg"] else None
            row[f"{region}_unit_price"] = _to_number(sheet.cell(row_idx, cols["unit"]).value) if cols["unit"] else None
        rows.append(row)

    CATALOG_CACHE = rows
    return CATALOG_CACHE


def summarize_master_df(df: List[dict]) -> Dict:
    if not df:
        return {"columns": [], "rows": [], "count": 0}

    preview_cols = [
        "code",
        "name",
        "spec",
        f"{REGION_SU}_spec_price",
        f"{REGION_SU}_kg_price",
        f"{REGION_SU}_unit_price",
        f"{REGION_GS}_spec_price",
        f"{REGION_GS}_unit_price",
        f"{REGION_GN}_spec_price",
        f"{REGION_GN}_unit_price",
        f"{REGION_HN}_spec_price",
        f"{REGION_HN}_unit_price",
    ]
    rows = [{column: row.get(column) for column in preview_cols} for row in df[:200]]
    return {"columns": preview_cols, "rows": rows, "count": len(df)}


def get_field_map(df: List[dict]) -> Dict[str, Optional[str]]:
    if not df:
        return {}
    return {
        "name": "name",
        "spec": "spec",
        f"{REGION_SU}_spec": f"{REGION_SU}_spec_price",
        f"{REGION_SU}_kg": f"{REGION_SU}_kg_price",
        f"{REGION_SU}_unit": f"{REGION_SU}_unit_price",
        f"{REGION_GS}_spec": f"{REGION_GS}_spec_price",
        f"{REGION_GS}_unit": f"{REGION_GS}_unit_price",
        f"{REGION_GN}_spec": f"{REGION_GN}_spec_price",
        f"{REGION_GN}_unit": f"{REGION_GN}_unit_price",
        f"{REGION_HN}_spec": f"{REGION_HN}_spec_price",
        f"{REGION_HN}_unit": f"{REGION_HN}_unit_price",
    }


def clean_product_name(name: str) -> str:
    return _safe_str(name).replace("\n", " ").strip()


def find_best_match(
    df: List[dict],
    query: str,
    spec: str = "",
    prices: Optional[Dict[str, Optional[int]]] = None,
    region: str = REGION_SU,
) -> Tuple[Optional[dict], float]:
    if not df:
        return None, 0.0

    query_candidates = name_candidates(query)
    if not query_candidates:
        return None, 0.0

    best_row: Optional[dict] = None
    best_score = 0.0

    for row in df:
        row_pool = list(row.get("name_candidates") or [])
        row_pool.append(row.get("normalized_name", ""))
        name_score = 0.0
        for query_candidate in query_candidates:
            for row_candidate in row_pool:
                if not query_candidate or not row_candidate:
                    continue
                score = float(
                    max(
                        fuzz.WRatio(query_candidate, row_candidate),
                        fuzz.token_sort_ratio(query_candidate, row_candidate),
                        _containment_bonus(query_candidate, row_candidate),
                    )
                )
                if score >= 95:
                    score += min(len(query_candidate), len(row_candidate)) * 0.2
                name_score = max(name_score, score)

        spec_score = _spec_similarity(spec, row.get("spec", "")) if spec else 0.0
        price_bonus = 0.0
        if prices:
            for kind, weight in (("spec_price", 8.0), ("kg_price", 6.0), ("unit_price", 6.0)):
                left = prices.get(kind)
                right = row.get(f"{region}_{kind}")
                if left is not None and right is not None and abs(int(left) - int(right)) <= 1:
                    price_bonus += weight

        total = round(name_score * 0.78 + spec_score * 0.22 + price_bonus, 2)
        if total > best_score:
            best_score = total
            best_row = row

    if not best_row or best_score < NAME_MATCH_THRESHOLD:
        return None, best_score
    return best_row, best_score


def extract_master_prices(row: dict, region: str, df: List[dict]) -> Dict[str, Optional[int]]:
    del df
    return {
        "spec_price": row.get(f"{region}_spec_price"),
        "kg_price": row.get(f"{region}_kg_price"),
        "unit_price": row.get(f"{region}_unit_price"),
        "spec_text": row.get("spec"),
        "product_name": row.get("name"),
    }
